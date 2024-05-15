from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

# Функция для получения кошельков из Excel-таблицы
def get_wallets_from_excel(file_path):
    wallets = []
    wb = load_workbook(filename=file_path)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        wallets.append(row[0])  # Предполагается, что кошельки находятся в первом столбце
    return wallets

# Открываем браузер
driver_path = r"C:\путь_к_вашему_файлу\chromedriver.exe"


  # Замените путь на путь к вашему драйверу браузера
driver = webdriver.Chrome(driver_path)

# Получаем кошельки из Excel-таблицы
wallets = get_wallets_from_excel("Metamask_10000.xlsx")

for wallet in wallets:
    # Открыть сайт Metamask
    driver.get("https://metamask.io/")
    time.sleep(5)  # Подождать 5 секунд

    # Добавить код для подключения кошелька Metamask
    # ...

    # Открыть сайт крана Holesky
    driver.get("https://holesky-faucet.gasp.xyz/")
    time.sleep(5)  # Подождать 5 секунд

    # Нажать на кнопку "Connected to Holesky"
    connected_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Connected to Holesky')]"))
    )
    connected_button.click()

    # Сменить сеть на Holesky
    # ...

    # Нажать на кнопки "CLAIM GETH" и "CLAIM GASP"
    # ...

    # Задержка перед закрытием браузера
    time.sleep(5)  # Подождать 5 секунд

    # Проверка на наличие окна браузера
    if driver.window_handles:
        # Закрыть браузер
        driver.quit()
